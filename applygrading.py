#!/usr/bin/python3
## Take a list of grading codes and update it in the Rubric-style grading sheet
## Author:  Joseph T. Foley <foley AT RU DOT IS>
## Start Date: 2022-09-12
## Input: Spreadsheet, List of Variables "Name=", penalty and credit codes e.g. "INT1(-2)" and "!MISC"
## In the future:  This list will be supplied by the PDF comment extractor
## Output:  Updated Spreadsheet
## Ubuntu install
##   sudo apt install python3-openpyxl
## Doc: https://openpyxl.readthedocs.io/en/stable/tutorial.html 
from openpyxl import load_workbook
wb = load_workbook("templates/notebook-eval.xlsx")
#print(wb.sheetnames)
mysheet = wb["team1"]

# STUB:  We know that codes are in column A
# STUB:  We know that we need to update values in column C
